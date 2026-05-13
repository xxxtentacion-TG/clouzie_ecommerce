import csv
import io
import json
import os
from datetime import timedelta
from django.conf import settings

from django.core.paginator import Paginator
from django.db.models import Count, F, Sum
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone

from adminpanel.templatetags.currency import format_inr
from adminpanel.utils.sales_analytics import (
    build_chart_data,
    calculate_filtered_metrics,
    filtered_revenue_orders,
    filtered_successful_orders,
    order_count_for_range,
    sales_for_previous_day,
)
from orders.models import OrderItem


FILTER_OPTIONS = [
    ("Daily", "daily"),
    ("Weekly", "weekly"),
    ("Monthly", "monthly"),
    ("Yearly", "yearly"),
]


def _percent_change(current, previous):
    if previous == 0:
        return 100 if current > 0 else 0
    return int(((current - previous) / previous) * 100)


def _report_filename(extension):
    return f"sales_report_{timezone.localdate().isoformat()}.{extension}"


def _orders_for_export(filter_type, start_str, end_str):
    return filtered_successful_orders(filter_type, start_str, end_str).order_by("-placed_at")


def sales_report(request):
    filter_type = request.GET.get("filter", "monthly")
    start_str = request.GET.get("start_date", "")
    end_str = request.GET.get("end_date", "")
    page_number = request.GET.get("page", 1)

    metrics = calculate_filtered_metrics(filter_type, start_str, end_str)
    revenue_orders = filtered_revenue_orders(filter_type, start_str, end_str)
    successful_orders = filtered_successful_orders(filter_type, start_str, end_str)

    today = timezone.localdate()
    yesterday = today - timedelta(days=1)
    today_revenue = sales_for_previous_day(today)
    yesterday_revenue = sales_for_previous_day(yesterday)

    week_start = today - timedelta(days=today.weekday())
    previous_week_start = week_start - timedelta(days=7)
    previous_week_end = week_start - timedelta(days=1)
    current_week_orders = order_count_for_range(week_start, today)
    previous_week_orders = order_count_for_range(previous_week_start, previous_week_end)

    revenue_change = _percent_change(today_revenue, yesterday_revenue)
    orders_change = _percent_change(current_week_orders, previous_week_orders)

    top_products = (
        OrderItem.objects.filter(order__in=revenue_orders)
        .exclude(status="CANCELLED")
        .values("product_name")
        .annotate(total_sold=Sum("quantity"), revenue=Sum("total"))
        .order_by("-total_sold")[:10]
    )

    top_categories_qs = (
        OrderItem.objects.filter(order__in=revenue_orders)
        .exclude(status="CANCELLED")
        .values(cat_name=F("variant__product__subcategory__category__name"))
        .annotate(total_orders=Count("id"), revenue=Sum("total"))
        .order_by("-revenue")[:6]
    )
    max_cat_rev = max((c["revenue"] or 0 for c in top_categories_qs), default=1)
    top_categories = [
        {**c, "pct": int((c["revenue"] or 0) / max_cat_rev * 100)}
        for c in top_categories_qs
    ]

    paginator = Paginator(successful_orders.order_by("-placed_at"), 10)
    orders_page = paginator.get_page(page_number)

    context = {
        "orders": orders_page,
        "filter_type": filter_type,
        "start_date": start_str,
        "end_date": end_str,
        "filter_options": FILTER_OPTIONS,
        "total_orders": metrics["total_orders"],
        "gross_revenue": metrics["gross_revenue"],
        "coupon_discount": metrics["coupon_discount"],
        "offer_discount": metrics["offer_discount"],
        "net_revenue": metrics["net_revenue"],
        "total_customers": metrics["total_customers"],
        "refund_amount": metrics["refund_amount"],
        "avg_order_value": metrics["avg_order_value"],
        "revenue_insight": f"Revenue {'up' if revenue_change >= 0 else 'down'} {abs(revenue_change)}% vs yesterday",
        "orders_insight": f"Orders {'up' if orders_change >= 0 else 'down'} {abs(orders_change)}% this week",
        "revenue_change": revenue_change,
        "orders_change": orders_change,
        "top_products": top_products,
        "top_categories": top_categories,
        "chart_data_json": json.dumps(build_chart_data()),
    }
    return render(request, "adminpanel/sales_report.html", context)


def export_sales_csv(request):
    filter_type = request.GET.get("filter", "daily")
    start_str = request.GET.get("start_date", "")
    end_str = request.GET.get("end_date", "")
    orders = _orders_for_export(filter_type, start_str, end_str)

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{_report_filename("csv")}"'

    writer = csv.writer(response)
    writer.writerow([
        "Order ID", "Customer", "Payment Method",
        "Coupon Discount", "Final Amount", "Payment Status", "Order Status", "Date",
    ])
    for order in orders:
        writer.writerow([
            order.order_id,
            order.user.username,
            order.get_payment_method_display(),
            format_inr(order.discount_amount),
            format_inr(order.total_amount),
            order.get_payment_status_display(),
            order.get_order_status_display(),
            timezone.localtime(order.placed_at).strftime("%d %b %Y"),
        ])
    return response


def export_sales_excel(request):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    filter_type = request.GET.get("filter", "daily")
    start_str = request.GET.get("start_date", "")
    end_str = request.GET.get("end_date", "")
    orders = _orders_for_export(filter_type, start_str, end_str)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    ws.append(["Order ID", "Date", "Customer", "Total Amount", "Payment Status", "Order Status"])

    header_fill = PatternFill("solid", fgColor="111111")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 20

    for order in orders:
        ws.append([
            order.order_id,
            timezone.localtime(order.placed_at).strftime("%d %b %Y"),
            order.user.username,
            format_inr(order.total_amount),
            order.get_payment_status_display(),
            order.get_order_status_display(),
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{_report_filename("xlsx")}"'
    return response


def export_sales_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    # Register Unicode fonts that support ₹ symbol
    font_dir = os.path.join(settings.BASE_DIR, "static", "fonts")
    pdfmetrics.registerFont(TTFont("Calibri",      os.path.join(font_dir, "Calibri.ttf")))
    pdfmetrics.registerFont(TTFont("Calibri-Bold", os.path.join(font_dir, "Calibri-Bold.ttf")))

    filter_type = request.GET.get("filter", "daily")
    start_str   = request.GET.get("start_date", "")
    end_str     = request.GET.get("end_date",   "")
    orders  = _orders_for_export(filter_type, start_str, end_str)
    metrics = calculate_filtered_metrics(filter_type, start_str, end_str)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=2*cm,   bottomMargin=2*cm,
    )
    elements = []

    # Title
    elements.append(Paragraph(
        "CLOUZIE — Sales Report",
        ParagraphStyle("title", fontName="Calibri-Bold", fontSize=18, spaceAfter=4),
    ))

    period_label = filter_type.capitalize()
    if filter_type == "custom" and start_str and end_str:
        period_label = f"{start_str} to {end_str}"
    elements.append(Paragraph(
        f"Period: {period_label} | Generated: {timezone.localtime().strftime('%d %b %Y, %H:%M')}",
        ParagraphStyle("sub", fontName="Calibri", fontSize=9, textColor=colors.grey, spaceAfter=12),
    ))

    # Summary table
    # format_inr() already returns "₹12,000" — do NOT prefix with ₹ again
    def rupee(val):
        return format_inr(val)  # single ₹ guaranteed

    summary_data = [
        ["Metric", "Amount"],
        ["Total Orders",     str(metrics["total_orders"])],
        ["Gross Revenue",    rupee(metrics["gross_revenue"])],
        ["Offer Discounts",  rupee(metrics["offer_discount"])],
        ["Coupon Discounts", rupee(metrics["coupon_discount"])],
        ["Refunds",          rupee(metrics["refund_amount"])],
        ["Net Revenue",      rupee(metrics["net_revenue"])],
        ["Avg Order Value",  rupee(metrics["avg_order_value"])],
    ]
    summary_tbl = Table(summary_data, colWidths=[6*cm, 5*cm])
    summary_tbl.setStyle(TableStyle([
        ("BACKGROUND",     (0,0), (-1,0), colors.black),
        ("TEXTCOLOR",      (0,0), (-1,0), colors.white),
        ("FONTNAME",       (0,0), (-1,0), "Calibri-Bold"),
        ("FONTNAME",       (0,1), (-1,-1), "Calibri"),
        ("FONTSIZE",       (0,0), (-1,-1), 9),
        ("GRID",           (0,0), (-1,-1), 0.5, colors.lightgrey),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f9f9f9")]),
        ("LEFTPADDING",    (0,0), (-1,-1), 8),
        ("RIGHTPADDING",   (0,0), (-1,-1), 8),
        ("TOPPADDING",     (0,0), (-1,-1), 5),
        ("BOTTOMPADDING",  (0,0), (-1,-1), 5),
    ]))
    elements.append(summary_tbl)
    elements.append(Spacer(1, 0.5*cm))

    # Orders table
    order_rows = [["Order ID", "Customer", "Method", "Coupon Off", "Amount", "Payment", "Status", "Date"]]
    for order in orders[:200]:
        order_rows.append([
            order.order_id,
            order.user.username[:20],
            order.get_payment_method_display(),
            rupee(order.discount_amount),
            rupee(order.total_amount),
            order.get_payment_status_display(),
            order.get_order_status_display(),
            timezone.localtime(order.placed_at).strftime("%d %b %Y"),
        ])

    order_tbl = Table(
        order_rows,
        colWidths=[3.5*cm, 3.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 3*cm, 2.5*cm],
        repeatRows=1,
    )
    order_tbl.setStyle(TableStyle([
        ("BACKGROUND",     (0,0), (-1,0), colors.black),
        ("TEXTCOLOR",      (0,0), (-1,0), colors.white),
        ("FONTNAME",       (0,0), (-1,0), "Calibri-Bold"),
        ("FONTNAME",       (0,1), (-1,-1), "Calibri"),
        ("FONTSIZE",       (0,0), (-1,-1), 8),
        ("GRID",           (0,0), (-1,-1), 0.4, colors.lightgrey),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f9f9f9")]),
        ("LEFTPADDING",    (0,0), (-1,-1), 5),
        ("RIGHTPADDING",   (0,0), (-1,-1), 5),
        ("TOPPADDING",     (0,0), (-1,-1), 4),
        ("BOTTOMPADDING",  (0,0), (-1,-1), 4),
    ]))
    elements.append(order_tbl)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{_report_filename("pdf")}"'
    return response
